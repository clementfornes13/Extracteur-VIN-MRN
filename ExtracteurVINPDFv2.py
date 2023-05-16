from os import (path, listdir)
from sys import (exit, argv)
from re import (findall)
from PyPDF2 import (PdfReader)
from openpyxl import (Workbook)
from PyQt6.QtGui import (QIcon, QPixmap, QDesktopServices)
from PyQt6.QtCore import (QCoreApplication, QUrl, Qt)
from PyQt6.QtWidgets import (QApplication, QMainWindow, QStyleOption,QStyle,QFileDialog, QLabel,QPushButton, QLineEdit, QVBoxLayout, QGridLayout, QWidget,QMessageBox, QProgressBar)

class VINExtractor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ListeVIN = []
        self.ListeMRN = []
        self.unique_VIN = []
        self.IMG_PATH = path.join(path.dirname(__file__), 'images')
        self.VIN_PATTERN = r"(?!FR)[A-HJ-NPR-Z][A-HJ-NPR-Z0-9]{15}[A-HJ-NPR-Z\d]"
        self.MRN_PATTERN = r"(?<!\S)(?!MRN)(?=[A-Z0-9]{18}\b)[A-Z0-9]*[A-Z][A-Z0-9]*\b(?!\S)"
        self.icon = path.join(self.IMG_PATH, 'Icone.ico').replace('\\','/')
        self.logo = path.join(self.IMG_PATH, 'Logo TEA FOS.png').replace('\\','/')
        self.initUI()
    def initUI(self):
        self.setWindowTitle("Extracteur VIN MRN PDF EAD")
        self.setWindowIcon(QIcon(self.icon))
        centralWidget = QWidget()
        self.setCentralWidget(centralWidget)
        mainLayout = QVBoxLayout()
        centralWidget.setLayout(mainLayout)
        self.resize(450,100)
        logoLabel = QLabel()
        logoPixmap = QPixmap(self.logo)
        logoLabel.setPixmap(logoPixmap)
        logoLabel.setAlignment(Qt.AlignmentFlag.AlignHCenter)
        mainLayout.addWidget(logoLabel)
        gridLayout = QGridLayout()
        mainLayout.addLayout(gridLayout)
        labelPDF = QLabel("Dossier PDF :")
        gridLayout.addWidget(labelPDF, 0, 0)
        self.inputPDF = QLineEdit()
        gridLayout.addWidget(self.inputPDF, 0, 1)
        browsePDF = QPushButton()
        browsePDF.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon, QStyleOption()))
        browsePDF.clicked.connect(self.browsePDFClicked)
        gridLayout.addWidget(browsePDF, 0, 2)
        labelDest = QLabel("Destination :")
        gridLayout.addWidget(labelDest, 1, 0)
        self.inputDest = QLineEdit()
        gridLayout.addWidget(self.inputDest, 1, 1)
        browseDest = QPushButton()
        browseDest.setIcon(QApplication.style().standardIcon(QStyle.StandardPixmap.SP_DirOpenIcon, QStyleOption()))
        browseDest.clicked.connect(self.browseDestClicked)
        gridLayout.addWidget(browseDest, 1, 2)
        extractButton = QPushButton("Extraction")
        extractButton.clicked.connect(self.launch_extraction)
        mainLayout.addWidget(extractButton)
        self.progress_bar = QProgressBar()
        mainLayout.addWidget(self.progress_bar)
        self.progress_bar.hide() # hide progress bar by default
    def browsePDFClicked(self):
        if directory := QFileDialog.getExistingDirectory(self, "Choisir un dossier"):
            self.inputPDF.setText(directory)
            return directory
    def browseDestClicked(self):
        if directory:=QFileDialog.getExistingDirectory(self, "Choisir un dossier"):
            self.inputDest.setText(directory)
            return directory
    def extract_vins_mrns(self, emplacement_pdf, destination=None):
        if not emplacement_pdf:
            return None
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Liste des VIN'
        sheet['B1'] = 'Liste des MRN'
        row_num = 2
        file_list = [pdf for pdf in listdir(emplacement_pdf) if pdf.lower().endswith('.pdf')]
        total_files = len(file_list)
        self.progress_bar.setMaximum(total_files)
        self.progress_bar.setValue(0)
        self.progress_bar.show()
        file_count = 0
        for file_count, pdf in enumerate(file_list, start=1):
            file_count += 1
            pdf = path.join(emplacement_pdf, pdf).replace("\\", "/")
            with open(pdf, 'rb') as file:
                lire_pdf = PdfReader(file)
                for page_num in range(len(lire_pdf.pages)):
                    page = lire_pdf.pages[page_num]
                    texte = page.extract_text()
                    vins = findall(self.VIN_PATTERN, texte)
                    mrns = findall(self.MRN_PATTERN, texte)
                    for vin in vins:
                            self.ListeVIN.append(vin)
                    for vin_unique in self.ListeVIN:
                        if vin_unique not in self.unique_VIN:
                            self.unique_VIN.append(vin_unique)
                    for mrn in mrns:
                        self.ListeMRN.append(mrn)
                    for y in range(len(self.unique_VIN)):
                        sheet.cell(row=row_num, column=1, value=self.unique_VIN[y])
                        sheet.cell(row=row_num, column=2, value=self.ListeMRN[0])
                        row_num += 1
                    self.ListeVIN=[]
                    self.ListeMRN=[]
                    self.unique_VIN = []
                file_name = "Extraction VIN MRN EAD.xlsx".format(1)
            self.progress_bar.setValue(file_count)
        if not destination:
            return None
        file_path = path.join(destination, file_name)
        i = 1
        while path.exists(file_path):
            i += 1
            file_name = f"Extraction VIN MRN EAD {i}.xlsx"
            file_path = path.join(destination, file_name)
        workbook.save(file_path)
        return file_path
    def launch_extraction(self, values):
        emplacement_pdf = self.inputPDF.text()
        destination = self.inputDest.text()
        file_list = [pdf for pdf in listdir(emplacement_pdf) if pdf.lower().endswith('.pdf')]
        total_files = len(file_list)
        if total_files==0:
            QMessageBox.critical(self,'Erreur', 'Le dossier PDF sélectionné est vide')
            return None
        file_path = self.extract_vins_mrns(emplacement_pdf, destination)
        if not file_path:
            QMessageBox.critical(self, 'Erreur', 'Pas de Dossier PDF / Destination choisi')
            return None
        QMessageBox.information(self, 'Extraction réussie', f"Fichier enregistré ici : {file_path}")
        QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
    def run(self):
        self.show()
        exit(QCoreApplication.instance().exec())
if __name__ == '__main__':
    app = QApplication(argv)
    extracteur = VINExtractor()
    extracteur.run()
    exit(app.exec())