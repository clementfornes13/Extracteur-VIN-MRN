##############################################################################
# Nom du fichier: Extracteur VIN MRN.py
# Description: Ce script permet d'extraire les VIN et MRN de fichiers PDF
#
# Auteur: FORNES Clément
# Date: 4 Juillet 2023
#
# Copyright (c) 2023 FORNES Clément
#
# Licence: MIT Licfacebookense
##############################################################################
# MIT License
#
# Copyright (c) 2023 FORNES Clément
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
#
# ------------------------------------------------------------------------------
#
# Ce script permet d'extraire les VIN et MRN de fichiers PDF
# Il est possible d'extraire les VIN, les MRN ou les deux
# Les VIN et MRN sont enregistrés dans un fichier Excel
# Le fichier Excel est enregistré dans le dossier de destination
# Le fichier Excel est ouvert automatiquement après l'extraction
#
# ------------------------------------------------------------------------------
# Mise à jour: 8 Septembre 2023
#################### MODULES ####################
from tkinter import Tk, Label, Button, Entry, filedialog, messagebox, PhotoImage, IntVar, END
from tkinter.ttk import Progressbar, Frame, Radiobutton
from openpyxl import Workbook
from PyPDF2 import PdfReader
from os import path, startfile, listdir
from re import findall
import threading
import datetime
#################################################

#################### CLASSES ####################
class VINMRNExtractor(Tk):

    def __init__(self):
        #################### VARIABLES ####################
        self.ListeVIN, self.ListeMRN, self.unique_VIN, self.unique_MRN = [], [], [], []
        self.IMG_PATH = path.join(path.dirname(__file__), 'images')
        # Patterne des VIN
        # Explication :
        # (?!FR) : Ne pas commencer par FR
        # [A-HJ-NPR-Z] : Lettre de A à Z sauf I, O et Q
        # Même chose pour les 15 caractères suivants
        # [A-HJ-NPR-Z\d] : Lettre de A à Z sauf I, O et Q ou chiffre
        # Cette expression régulière
        self.VIN_PATTERN = r"(?!FR)[A-HJ-NPR-Z][A-HJ-NPR-Z0-9]{15}[A-HJ-NPR-Z\d]"
        # Patterne des MRN
        # Explication :
        # (?<!\S) : Ne pas commencer par un caractère
        # (?!MRN) : Ne pas commencer par MRN
        # (?=[A-Z0-9]{18}\b) : Commencer par 18 caractères alphanumériques
        # [A-Z0-9]*[A-Z][A-Z0-9]*\b : Finir par un caractère alphanumérique
        # (?!\S) : Ne pas finir par un caractère
        self.MRN_PATTERN = r"(?<!\S)(?!MRN)(?=[A-Z0-9]{18}\b)[A-Z0-9]*[A-Z][A-Z0-9]*\b(?!\S)"
        #################################################
        # Initialisation de la fenêtre principale
        super().__init__()
        self.icon = path.join(self.IMG_PATH, 'Icone.ico').replace('\\', '/')
        self.logo = path.join(self.IMG_PATH, 'Logo TEA FOS.png').replace('\\', '/')
        self.initUI()

    def initUI(self):
        # Initialisation de l'interface graphique
        # Paramètres de la fenêtre
        self.title("Extracteur VIN MRN")
        self.iconbitmap(self.icon)
        self.geometry("350x320")
        logoLabel = Label(self)
        logoImage = PhotoImage(file=self.logo)
        logoLabel.config(image=logoImage)
        logoLabel.image = logoImage
        logoLabel.grid(row=0, column=0, pady=10)
        # Frame
        frame = Frame(self)
        frame.grid(row=1, column=0)
        # Dossier PDF
        labelPDF = Label(frame, text="Dossier PDF:")
        labelPDF.grid(row=0, column=0, padx=5, pady=5)
        self.inputPDF = Entry(frame, width=30)
        self.inputPDF.grid(row=0, column=1, padx=5, pady=5)
        browsePDF = Button(frame, text="Parcourir", command=self.browsePDFClicked)
        browsePDF.grid(row=0, column=2, padx=5, pady=5)
        # Destination
        labelDest = Label(frame, text="Destination:")
        labelDest.grid(row=1, column=0, padx=5, pady=5)
        self.inputDest = Entry(frame, width=30)
        self.inputDest.grid(row=1, column=1, padx=5, pady=5)
        browseDest = Button(frame, text="Parcourir", command=self.browseDestClicked)
        browseDest.grid(row=1, column=2, padx=5, pady=5)
        # Boutons
        self.extractButton = Button(self, text="Extraction", command=self.launch_extraction)
        self.extractButton.grid(row=5, column=0, pady=10)
        # Options d'extraction
        self.extractOption = IntVar()
        self.extractOption.set(4)
        # 1 = Extraction VIN
        self.vinRadio = Radiobutton(self, text="VIN", variable=self.extractOption, value=1)
        self.vinRadio.grid(row=2, column=0, padx=5, pady=5, sticky='n')
        # 2 = Extraction MRN
        self.mrnRadio = Radiobutton(self, text="MRN", variable=self.extractOption, value=2)
        self.mrnRadio.grid(row=3, column=0, padx=5, pady=5, sticky='n')
        # 3 = Extraction VIN et MRN
        self.vinMrnRadio = Radiobutton(self, text="VIN + MRN", variable=self.extractOption, value=3)
        self.vinMrnRadio.grid(row=4, column=0, padx=5, pady=5, sticky='n')
        # Barre de progression
        self.progress_bar = Progressbar(self, mode='determinate')
        self.progress_bar.grid(row=6, column=0, pady=5)
        self.progress_bar.grid_remove()
        # Crée un bouton d'aide personnalisé
        help_button = Button(self, text="Aide", command=self.show_help)
        help_button.grid(row=0, column=0,pady=5,sticky='ne')

    # Définit la fonction pour afficher l'aide
    def show_help(self):
        help_text = (
            "Bienvenue dans l'Extracteur VIN MRN !\n\n"
            + "Utilisez ce programme pour extraire les VIN et MRN à partir de fichiers PDF.\n\n"
        )
        help_text += "Instructions :\n"
        help_text += "1. Cliquez sur le bouton 'Parcourir' à côté de 'Dossier PDF' pour sélectionner le dossier contenant les fichiers PDF à traiter.\n"
        help_text += "2. Cliquez sur le bouton 'Parcourir' à côté de 'Destination' pour choisir où enregistrer les résultats.\n"
        help_text += "3. Sélectionnez l'option d'extraction souhaitée : 'VIN', 'MRN' ou 'VIN + MRN'.\n"
        help_text += "4. Cliquez sur le bouton 'Extraction' pour démarrer le processus d'extraction.\n"
        help_text += "5. La barre de progression indiquera l'avancement de l'extraction.\n"
        help_text += "6. Les résultats seront enregistrés dans le dossier de destination que vous avez choisi.\n"
        help_text += "\nN'oubliez pas de vous assurer que les fichiers PDF contiennent les informations VIN et MRN que vous souhaitez extraire.\n"
        help_text += "\nSi vous rencontrez des problèmes, veuillez contacter FORNES Clément à l'adresse suivante : clement.fornes@teamarseille.gcatrans.com"
        messagebox.showinfo("Aide", help_text)

    # Désactiver les boutons
    def disable_buttons(self):
        self.enable_disable_buttons("disabled")
    # Activer les boutons
    def enable_buttons(self):
        self.enable_disable_buttons("normal")

    def enable_disable_buttons(self, state):
        self.extractButton.config(state=state)
        self.vinRadio.config(state=state)
        self.mrnRadio.config(state=state)
        self.vinMrnRadio.config(state=state)
    # Fenêtre de sélection du dossier PDF
    def browsePDFClicked(self):
        # Selection du dossier contenant les PDF
        if directory := filedialog.askdirectory(title="Choisir un dossier"):
            self.inputPDF.delete(0, END)
            self.inputPDF.insert(0, directory)

    # Fenêtre de sélection du dossier de destination
    def browseDestClicked(self):
        # Selection du dossier de destination
        if directory := filedialog.askdirectory(title="Choisir un dossier"):
            self.inputDest.delete(0, END)
            self.inputDest.insert(0, directory)

    # Extraction des VIN et MRN
    def extract_vins_mrns(self, emplacement_pdf, destination, extract_option):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = 'Liste des VIN'
        sheet['B1'] = 'Liste des MRN'
        row_num = 2
        file_list = [pdf for pdf in listdir(emplacement_pdf) if pdf.lower().endswith('.pdf')]
        total_files = len(file_list)
        self.progress_bar['maximum'] = total_files
        self.progress_bar.grid()
        for file_count, pdf in enumerate(file_list, start=1):
            pdf = path.join(emplacement_pdf, pdf).replace("\\", "/")
            with open(pdf, 'rb') as file:
                lire_pdf = PdfReader(file)
                if extract_option == 1:
                    # Pour chaque page du PDF
                    for _, page in enumerate(lire_pdf.pages):
                        texte = page.extract_text()
                        vins = findall(self.VIN_PATTERN, texte)
                        self.ListeVIN.extend(vins)
                        # Avoir une liste de VIN uniques
                        for vin_unique in set(self.ListeVIN):
                            self.unique_VIN.append(vin_unique)
                        # Ajoute chaque VIN par ligne sur la colonne A
                        for vin in self.unique_VIN:
                            sheet.cell(row=row_num, column=1, value=vin)
                            row_num += 1
                        self.ListeVIN, self.unique_VIN = [], []
                elif extract_option == 2:
                    for _, page in enumerate(lire_pdf.pages):
                        texte = page.extract_text()
                        mrns = findall(self.MRN_PATTERN, texte)
                        self.ListeMRN.extend(mrns)
                        for mrn_unique in set(self.ListeMRN):
                            self.unique_MRN.append(mrn_unique)
                        for mrn in self.unique_MRN:
                            sheet.cell(row=row_num, column=1, value=mrn)
                            row_num += 1
                        self.ListeMRN, self.unique_MRN = [], []
                elif extract_option == 3:
                    for _, page in enumerate(lire_pdf.pages):
                        texte = page.extract_text()
                        vins = findall(self.VIN_PATTERN, texte)
                        mrns = findall(self.MRN_PATTERN, texte)
                        self.ListeVIN.extend(vins)
                        self.ListeMRN.extend(mrns)
                        for vin_unique in set(self.ListeVIN):
                            self.unique_VIN.append(vin_unique)
                        for vin in self.unique_VIN:
                            sheet.cell(row=row_num, column=1, value=vin)
                            if len(self.ListeMRN) > 0:
                                sheet.cell(row=row_num, column=2, value=self.ListeMRN[0])
                            row_num += 1
                        self.ListeVIN, self.ListeMRN, self.unique_VIN = [], [], []
            file_count += 1
            self.progress_bar['value'] = file_count
            self.update()
        if not destination:
            return None
        if extract_option == 1:
            return self.save_excel(
                'Extraction VIN EAD ', destination, workbook
            )
        elif extract_option == 2:
            return self.save_excel(
                'Extraction MRN EAD ', destination, workbook
            )
        elif extract_option == 3:
            return self.save_excel(
                'Extraction VIN MRN EAD ', destination, workbook
            )

    # Mise à jour de la barre de progression
    def update_progress(self, value, maximum):
        self.progress_bar['value'] = value
        self.progress_bar['maximum'] = maximum
        self.update()

    # Chemin de fichier temporaire
    def extraction_temp(self, extraction_func, args, file_path_temp):
        file_path_temp[0] = extraction_func(*args)

    # Message d'erreur
    def error_message(self, arg0):
        messagebox.showerror('Erreur', arg0)
        self.enable_buttons()
        return None

    # Enregistrement du fichier Excel
    def save_excel(self, arg0, destination, workbook):
        current_datetime = datetime.datetime.now()
        formatted_datetime = current_datetime.strftime("%d-%m-%Y %H-%M")
        file_name = f"{arg0}{formatted_datetime}.xlsx"
        file_path = path.join(destination, file_name)
        workbook.save(file_path)
        return file_path

    # Lancement de l'extraction
    def launch_extraction(self):
        self.disable_buttons()
        file_path_temp = [None]
        emplacement_pdf = self.inputPDF.get()
        destination = self.inputDest.get()
        if emplacement_pdf == '' or destination == '':
            return self.error_message(
                'Veuillez sélectionner un dossier source et un dossier de destination'
            )
        file_list = [pdf for pdf in listdir(emplacement_pdf) if pdf.lower().endswith('.pdf')]
        total_files = len(file_list)
        if total_files == 0:
            return self.error_message(
                'Le dossier source sélectionné est vide'
            )
        extract_option = self.extractOption.get()
        if extract_option in [1, 2, 3]:
            thread = threading.Thread(
                target=self.extraction_temp,
                args=(self.extract_vins_mrns, (
                    emplacement_pdf,
                    destination,
                    extract_option,
                    ),
                    file_path_temp
                ),
            )
        else:
            return self.error_message(
                'Veuillez sélectionner une option d\'extraction'
            )
        thread.start()
        self.progress_bar.grid()
        stop_flag = False
        while thread.is_alive() or threading.active_count() > 1:
            if not thread.is_alive() and threading.active_count() == 1:
                stop_flag = True
            self.update_progress(threading.active_count() - 1, total_files)
            if stop_flag:
                break
        self.progress_bar.grid_remove()
        for t in threading.enumerate():
            if t != threading.current_thread():
                t.join()
        file_path = file_path_temp[0]
        if not file_path:
            return self.error_message(
                'Une erreur est survenue lors de l\'extraction des données'
            )
        messagebox.showinfo('Extraction réussie', f"Fichier enregistré ici : {file_path}")
        try:
            startfile(file_path)
        except Exception:
            messagebox.showerror('Erreur', 'Impossible d\'ouvrir le fichier Excel')
        self.progress_bar.grid_remove()
        self.enable_buttons()


app = VINMRNExtractor()
app.resizable(False, False)
app.mainloop()
